                                        #### CZYTAJ ZANIM URUCHOMISZ!!!

        #Pliki i foldery konieczne do działania skryptu:

        # 1. folder 'rogi' z czterema jpegami przezentującymi rogi,
        # 2. folder 'excel' z czteremea plikami dla każdego rogu z koordynatami map
        # 3. foldery do zapisu map wyciętych i całych (domyslnie 'cale1', 'wyc1')
        # 4. pusta geobaza do zapisu map zgeoreferowanych (domyslnie 'georeferencja.gdb')

        # Skrypt testowano z pomocą pythona 3.9 przy użyciu aplikacji pycharm.

        ###### WAŻNE:   ########
        # Konieczne do działania skryptu jest posiadanie biblioteki arcpy, która dostępna jest tylko w przypadku posiadania zainstalowanego pakietu arcgis!
        # Biblioteki (cv2,shutil,openpyxl) należy niestety instalować w interpreterze arcgisowym co jest dość upierdliwe.

import arcpy
from openpyxl import load_workbook
import cv2
import requests
import shutil
import os

#funkcja dodaje zera do linku - konieczna przy pobieraniu map z bazy pig
def przelicznik(numer):
    if numer in range(1,10):
        return '000'
    elif numer in range(10,100):
        return '00'
    elif numer in range(100,1000):
        return '0'
    return ''


numer_ark1 = [i for i in range(500,502)] #wybór numeru arkusza
numer = [str(x) for x in numer_ark1]

#wlasciwy kod - pętla dla każdej z wybranych map
for numer_ark in numer:
    try:
        zera = przelicznik(int(numer_ark))
        print(zera)
        url = "http://bazadata.pgi.gov.pl/data/smgp/arkusze_skany/smgp" + zera + numer_ark + ".jpg"
        path = 'cale1/' + numer_ark + '.jpg'
        r = requests.get(url, stream=True)
        if r.status_code == 200:
            with open(path, 'wb') as f:
                r.raw.decode_content = True
                shutil.copyfileobj(r.raw, f)
                print(f'Arkusz numer {numer_ark} pobrany')

            #okreslenie sciezek wejsciowych i docelowych
        cwd = os.getcwd()
        cwd1 = os.path.join(cwd,"Georeferencja.gdb")
        path1 = os.path.join(cwd, path)
        a = os.path.exists(path1)
        if a == False:
            print("arkusza {} nie ma w katalogu".format(numer_ark))
            continue

        image = cv2.imread(path)

        corner_names = {
            "lt": (1,1),
            "rt": (0, 1),
            "lb": (1, 0),
            "rb": (0, 0)
        }

        corner_cords = []
        large_image = cv2.imread(path)

        for corner, flags in corner_names.items():
            path1 = os.path.join(cwd,"rogi" ,(corner + "rog.jpg"))
            method = cv2.TM_SQDIFF_NORMED # one of six possible methods

            # Read the images from the file
            small_image = cv2.imread(path1)

            result = cv2.matchTemplate(small_image, large_image, method)
            # We want the minimum squared difference
            mn,_,mnLoc,_ = cv2.minMaxLoc(result)

            # Draw the rectangle:
            # Extract the coordinates of our best match
            MPx,MPy = mnLoc

            # Step 2: Get the size of the template. This is the same size as the match.
            trows,tcols = small_image.shape[:2]

            x_mult, y_mult = flags
            temp_corner = (MPx+tcols*x_mult, MPy+trows*y_mult)
            corner_cords.append(temp_corner)


        lt, rt, lb, rb = corner_cords

        min_x, min_y = lt
        max_x, max_y = rb

        cropped_image = large_image[min_y:max_y, min_x:max_x]

        w = max_x - min_x
        h = max_y - min_y
        cv2.imwrite(cwd + "/wyc1/"+ numer_ark + 'cut.png', cropped_image)


        #szukanie współrzędnych wrzuconego rastra


        Dol = arcpy.management.GetRasterProperties("wyc/" + str(numer_ark)+ "cut.png", "BOTTOM", '')
        Prawy = arcpy.management.GetRasterProperties("wyc/" + str(numer_ark) + "cut.png", "RIGHT", '')
        Lewy = arcpy.management.GetRasterProperties("wyc/" + str(numer_ark)+ "cut.png", "LEFT", '')
        Gora = arcpy.management.GetRasterProperties("wyc/" + str(numer_ark)+ "cut.png", "TOP", '')

        #print(Dol,Prawy,Lewy,Gora)
        #szukanie wspolrzędnych docelowych rastra

        corner_cords = []
        large_image = cv2.imread(path)
        wspol = ["LG", "LD", "PG", "PD"]
        wsplx = {}
        wsply = {}

        # rogi docelowe

        for d in wspol:
            f = wspol.index(d)
            wblg = os.path.join(cwd, "excel", "{}_excel.xlsx".format(wspol[f]))
            wb = load_workbook(wblg)
            ws = wb.active
            rows = ws.iter_rows(max_row=1100, max_col=3)
            nr = []
            dl = []
            szer = []
            for a, b, c in rows:
                nr.append(a.value)
                dl.append(b.value)
                szer.append(c.value)

            for i in nr:
                if i == int(numer_ark):
                    a = nr.index(i)

            wsplx["x" + d] = dl[a]
            wsply["y" + d] = szer[a]

        for key, val in wsplx.items():
            exec(key + '=val')
        for key, val in wsply.items():
            exec(key + '=val')

        xPG = xPG +20
        def Model():  # Funkcja stworzona przy pomocy modelbuildera służąca do georeferencji

            #punkty wejsciowe
            sourceA = "{0} {1}".format(Lewy, Gora)
            sourceB = "{0} {1}".format(Lewy, Dol)
            sourceC = "{0} {1}".format(Prawy, Gora)
            sourceD = "{0} {1}".format(Prawy, Dol)

            #punkty docelowe
            targetA = "{0} {1}".format(xLG, yLG)
            targetB = "{0} {1}".format(xLD, yLD)
            targetC = "{0} {1}".format(xPG, yPG)
            targetD = "{0} {1}".format(xPD, yPD)
            # To allow overwriting outputs change overwriteOutput option to True.
            arcpy.env.overwriteOutput = True
            _cut_png_2_ = arcpy.Raster(cwd + "/wyc/"+ numer_ark + 'cut.png')

            # Process: Build Pyramids And Statistics (Build Pyramids And Statistics) (management)
            _cut_png = arcpy.management.BuildPyramidsandStatistics(in_workspace=_cut_png_2_, include_subdirectories="INCLUDE_SUBDIRECTORIES", build_pyramids="BUILD_PYRAMIDS", calculate_statistics="CALCULATE_STATISTICS", BUILD_ON_SOURCE="NONE", block_field="", estimate_statistics="NONE", x_skip_factor=1, y_skip_factor=1, ignore_values=[], pyramid_level=-1, SKIP_FIRST="NONE", resample_technique="NEAREST", compression_type="NONE", compression_quality=100, skip_existing="SKIP_EXISTING", where_clause="", sips_mode="NONE")[0]

            # Process: Project Raster (Project Raster) (management)
            c_ProjectRaster = cwd1 + "\\c{}_ProjectRaster".format(numer_ark)
            arcpy.management.ProjectRaster(in_raster=_cut_png_2_, out_raster=c_ProjectRaster, out_coor_system="PROJCS[\"ETRS_1989_Poland_CS92\",GEOGCS[\"GCS_ETRS_1989\",DATUM[\"D_ETRS_1989\",SPHEROID[\"GRS_1980\",6378137.0,298.257222101]],PRIMEM[\"Greenwich\",0.0],UNIT[\"Degree\",0.0174532925199433]],PROJECTION[\"Transverse_Mercator\"],PARAMETER[\"False_Easting\",500000.0],PARAMETER[\"False_Northing\",-5300000.0],PARAMETER[\"Central_Meridian\",19.0],PARAMETER[\"Scale_Factor\",0.9993],PARAMETER[\"Latitude_Of_Origin\",0.0],UNIT[\"Meter\",1.0]]", resampling_type="NEAREST", cell_size="1 1", geographic_transform=[], Registration_Point="", in_coor_system="PROJCS[\"ETRS_1989_Poland_CS92\",GEOGCS[\"GCS_ETRS_1989\",DATUM[\"D_ETRS_1989\",SPHEROID[\"GRS_1980\",6378137.0,298.257222101]],PRIMEM[\"Greenwich\",0.0],UNIT[\"Degree\",0.0174532925199433]],PROJECTION[\"Transverse_Mercator\"],PARAMETER[\"False_Easting\",500000.0],PARAMETER[\"False_Northing\",-5300000.0],PARAMETER[\"Central_Meridian\",19.0],PARAMETER[\"Scale_Factor\",0.9993],PARAMETER[\"Latitude_Of_Origin\",0.0],UNIT[\"Meter\",1.0]]", vertical="NO_VERTICAL")
            c_ProjectRaster = arcpy.Raster(c_ProjectRaster)

            # Process: Warp (Warp) (management)
            c_ProjectRaster_Warp = cwd1 +"\\c{}Warp".format(numer_ark)
            arcpy.management.Warp(in_raster=c_ProjectRaster, source_control_points=[sourceA, sourceB, sourceC, sourceD], target_control_points=[targetA, targetB, targetC, targetD], out_raster=c_ProjectRaster_Warp, transformation_type="SPLINE", resampling_type="NEAREST")
            c20_ProjectRaster_Warp = arcpy.Raster(c_ProjectRaster_Warp)
            arcpy.Delete_management(c_ProjectRaster)
        if __name__ == '__main__':
            # Global Environment settings
            with arcpy.EnvManager(scratchWorkspace=path1, workspace=path1):
                Model()
    except: #wypluwanie ewentualnych błędów.
         print("Nie udało się wygenreować poprawnie arkusza {}".format(numer_ark))
         continue

