import arcpy
from openpyxl import load_workbook
import cv2
import os


numer_ark1 = [*range(1,100,1)] #wybór numeru arkusza
numer = [str(x) for x in numer_ark1]
for numer_ark in numer:
    try:
            #okreslenie sciezek wejsciowych i docelowych
        cwd = os.getcwd()
        cwd1 = os.path.join(cwd,"georeferencja.gdb")
        path = os.path.join(cwd,"SMGPPNG", "smgp" +numer_ark)+".png"
        a = os.path.exists(path)
        if a == False:
            print("arkusza {} nie ma w katalogu".format(numer_ark))
            continue

        image = cv2.imread(path)

        corner_names = {
            "lt": (1,1),
            "rt": (0, 1),
            "lb": (1, 0),
            "rb": (0, 0)
        }

        corner_cords = []
        large_image = cv2.imread(path)

        for corner, flags in corner_names.items():
            path1 = os.path.join(cwd,"rogi" ,(corner + "rog.png"))
            method = cv2.TM_SQDIFF_NORMED # one of six possible methods

            # Read the images from the file
            small_image = cv2.imread(path1)

            result = cv2.matchTemplate(small_image, large_image, method)
            # We want the minimum squared difference
            mn,_,mnLoc,_ = cv2.minMaxLoc(result)

            # Draw the rectangle:
            # Extract the coordinates of our best match
            MPx,MPy = mnLoc

            # Step 2: Get the size of the template. This is the same size as the match.
            trows,tcols = small_image.shape[:2]

            x_mult, y_mult = flags
            temp_corner = (MPx+tcols*x_mult, MPy+trows*y_mult)
            corner_cords.append(temp_corner)


        lt, rt, lb, rb = corner_cords

        min_x, min_y = lt
        max_x, max_y = rb

        cropped_image = large_image[min_y:max_y, min_x:max_x]

        w = max_x - min_x
        h = max_y - min_y
        cv2.imwrite(cwd + "/"+ numer_ark + 'cut.png', cropped_image)


        #szukanie współrzędnych wrzuconego rastra


        Dol=arcpy.management.GetRasterProperties(str(numer_ark)+ "cut.png", "BOTTOM", '')
        Prawy=arcpy.management.GetRasterProperties(str(numer_ark) + "cut.png", "RIGHT", '')
        Lewy=arcpy.management.GetRasterProperties(str(numer_ark)+ "cut.png", "LEFT", '')
        Gora=arcpy.management.GetRasterProperties(str(numer_ark)+ "cut.png", "TOP", '')

        #szukanie wspolrzędnych docelowych rastra

        corner_cords = []
        large_image = cv2.imread(path)
        wspol = ["LG", "LD", "PG", "PD"]
        wsplx = {}
        wsply = {}

        # rogi docelowe

        for d in wspol:
                f = wspol.index(d)
                wblg = os.path.join(cwd, "excel", "{}_excel.xlsx".format(wspol[f]))
                wb = load_workbook(wblg)
                ws = wb.active
                rows = ws.iter_rows(max_row=1100, max_col=3)
                nr = []
                dl = []
                szer = []
                for a, b, c in rows:
                    nr.append(a.value)
                    dl.append(b.value)
                    szer.append(c.value)

                for i in nr:
                    if i == int(numer_ark):
                        a = nr.index(i)

                wsplx["x" + d] = dl[a]
                wsply["y" + d] = szer[a]
        for key, val in wsplx.items():
            exec(key + '=val')
        for key, val in wsply.items():
            exec(key + '=val')

        xPG = xPG +20
        def Model():  # Model

            #punkty wejsciowe
            sourceA = "{0} {1}".format(Lewy, Gora)
            sourceB = "{0} {1}".format(Lewy, Dol)
            sourceC = "{0} {1}".format(Prawy, Gora)
            sourceD = "{0} {1}".format(Prawy, Dol)

            #punkty docelowe
            targetA = "{0} {1}".format(xLG, yLG)
            targetB = "{0} {1}".format(xLD, yLD)
            targetC = "{0} {1}".format(xPG, yPG)
            targetD = "{0} {1}".format(xPD, yPD)
            # To allow overwriting outputs change overwriteOutput option to True.
            arcpy.env.overwriteOutput = True
            _cut_png_2_ = arcpy.Raster(cwd + "\\{}cut.png".format(numer_ark))

            # Process: Build Pyramids And Statistics (Build Pyramids And Statistics) (management)
            _cut_png = arcpy.management.BuildPyramidsandStatistics(in_workspace=_cut_png_2_, include_subdirectories="INCLUDE_SUBDIRECTORIES", build_pyramids="BUILD_PYRAMIDS", calculate_statistics="CALCULATE_STATISTICS", BUILD_ON_SOURCE="NONE", block_field="", estimate_statistics="NONE", x_skip_factor=1, y_skip_factor=1, ignore_values=[], pyramid_level=-1, SKIP_FIRST="NONE", resample_technique="NEAREST", compression_type="NONE", compression_quality=100, skip_existing="SKIP_EXISTING", where_clause="", sips_mode="NONE")[0]

            # Process: Project Raster (Project Raster) (management)
            c_ProjectRaster = cwd1 + "\\c{}_ProjectRaster".format(numer_ark)
            arcpy.management.ProjectRaster(in_raster=_cut_png_2_, out_raster=c_ProjectRaster, out_coor_system="PROJCS[\"ETRS_1989_Poland_CS92\",GEOGCS[\"GCS_ETRS_1989\",DATUM[\"D_ETRS_1989\",SPHEROID[\"GRS_1980\",6378137.0,298.257222101]],PRIMEM[\"Greenwich\",0.0],UNIT[\"Degree\",0.0174532925199433]],PROJECTION[\"Transverse_Mercator\"],PARAMETER[\"False_Easting\",500000.0],PARAMETER[\"False_Northing\",-5300000.0],PARAMETER[\"Central_Meridian\",19.0],PARAMETER[\"Scale_Factor\",0.9993],PARAMETER[\"Latitude_Of_Origin\",0.0],UNIT[\"Meter\",1.0]]", resampling_type="NEAREST", cell_size="1 1", geographic_transform=[], Registration_Point="", in_coor_system="PROJCS[\"ETRS_1989_Poland_CS92\",GEOGCS[\"GCS_ETRS_1989\",DATUM[\"D_ETRS_1989\",SPHEROID[\"GRS_1980\",6378137.0,298.257222101]],PRIMEM[\"Greenwich\",0.0],UNIT[\"Degree\",0.0174532925199433]],PROJECTION[\"Transverse_Mercator\"],PARAMETER[\"False_Easting\",500000.0],PARAMETER[\"False_Northing\",-5300000.0],PARAMETER[\"Central_Meridian\",19.0],PARAMETER[\"Scale_Factor\",0.9993],PARAMETER[\"Latitude_Of_Origin\",0.0],UNIT[\"Meter\",1.0]]", vertical="NO_VERTICAL")
            c_ProjectRaster = arcpy.Raster(c_ProjectRaster)

            # Process: Warp (Warp) (management)
            c_ProjectRaster_Warp = cwd1 +"\\c{}Warp".format(numer_ark)
            arcpy.management.Warp(in_raster=c_ProjectRaster, source_control_points=[sourceA, sourceB, sourceC, sourceD], target_control_points=[targetA, targetB, targetC, targetD], out_raster=c_ProjectRaster_Warp, transformation_type="SPLINE", resampling_type="NEAREST")
            c20_ProjectRaster_Warp = arcpy.Raster(c_ProjectRaster_Warp)
            arcpy.Delete_management(c_ProjectRaster)
        if __name__ == '__main__':
            # Global Environment settings
            with arcpy.EnvManager(scratchWorkspace=path1, workspace=path1):
                Model()
        print("dla arkusza {} gotowe".format(numer_ark))
    except:
        print("Zla tabela w arkuszu {}".format(numer_ark))
        continue

